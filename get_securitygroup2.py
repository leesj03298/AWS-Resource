import boto3
from itertools import zip_longest
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ===== 설정 =====
PROFILE = "default"
REGION  = "ap-northeast-2"
OUTFILE = "security_groups_centered.xlsx"
# ===============

session = boto3.Session(profile_name=PROFILE) if PROFILE else boto3.Session()
ec2 = session.client("ec2", region_name=REGION)

# 1. 기초 데이터 로드
vpc_map = {v['VpcId']: next((t['Value'] for t in v.get('Tags', []) if t['Key'] == 'Name'), v['VpcId']) 
           for v in ec2.describe_vpcs()['Vpcs']}
sgs = sorted(ec2.describe_security_groups()['SecurityGroups'], key=lambda x: x.get('GroupName', '').lower())
sg_names = {sg['GroupId']: sg.get('GroupName', sg['GroupId']) for sg in sgs}

# 2. 규칙 추출 함수 (KeyError 방지 및 로직 최적화)
def get_rule_list(perms):
    rules = []
    for p in perms:
        proto = "all" if p.get('IpProtocol') == "-1" else str(p.get('IpProtocol', '-'))
        
        # 포트 안전하게 가져오기 (KeyError 해결)
        f_port, t_port = p.get('FromPort'), p.get('ToPort')
        port = "-" if f_port is None else (str(f_port) if f_port == t_port else f"{f_port}-{t_port}")
        
        # 소스 통합 (IP, SG, PL)
        srcs = [(r.get('CidrIp') or r.get('CidrIpv6'), r.get('Description', '-')) for r in p.get('IpRanges', []) + p.get('Ipv6Ranges', [])]
        srcs += [(f"{g['GroupId']}({sg_names.get(g['GroupId'], g['GroupId'])})", g.get('Description', '-')) for g in p.get('UserIdGroupPairs', [])]
        srcs += [(pl['PrefixListId'], pl.get('Description', '-')) for pl in p.get('PrefixListIds', [])]
        
        for src, desc in (srcs or [("-", "-")]):
            rules.append([proto, port, src, desc])
    return rules or [["-", "-", "-", "-"]]

# 3. 데이터 구성
final_data, merge_info = [], []
current_row = 3

for sg in sgs:
    in_rules, out_rules = get_rule_list(sg.get('IpPermissions', [])), get_rule_list(sg.get('IpPermissionsEgress', []))
    vpc_info = [vpc_map.get(sg.get('VpcId'), 'NO_VPC'), sg.get('GroupName', ''), sg['GroupId']]
    
    sg_rows = 0
    for i, o in zip_longest(in_rules, out_rules, fillvalue=["-", "-", "-", "-"]):
        final_data.append(vpc_info + i + o)
        sg_rows += 1
    
    if sg_rows > 1:
        merge_info.append((current_row, current_row + sg_rows - 1))
    current_row += sg_rows

# 4. 엑셀 생성 및 스타일 적용
wb = Workbook()
ws = wb.active
ws.title = "SecurityGroups"

# 공통 스타일 정의
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
header_fill = PatternFill("solid", fgColor="E6E6E6")
header_font = Font(bold=True)

# 헤더 작성
headers = [("A1:A2", "VPC Name"), ("B1:B2", "SG Name"), ("C1:C2", "Group ID"), ("D1:F1", "Inbound Rule"), ("G1:G2", "비고(In)"), ("H1:J1", "Outbound Rule"), ("K1:K2", "비고(Out)")]
for r, t in headers:
    ws.merge_cells(r)
    cell = ws[r.split(':')[0]]
    cell.value, cell.fill, cell.font, cell.alignment, cell.border = t, header_fill, header_font, center_align, thin_border
    # 병합된 셀 테두리 적용
    for row in ws[r]:
        for c in row: c.border = thin_border

for addr, text in {"D2":"Type", "E2":"Port", "F2":"Source", "H2":"Type", "I2":"Port", "J2":"Source"}.items():
    c = ws[addr]
    c.value, c.fill, c.font, c.alignment, c.border = text, header_fill, header_font, center_align, thin_border

# 데이터 작성 및 정렬 적용
for r_idx, row_data in enumerate(final_data, start=3):
    for c_idx, val in enumerate(row_data, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.alignment, cell.border = center_align, thin_border

# 세로 병합 (A, B, C열)
for start, end in merge_info:
    for col in "ABC":
        ws.merge_cells(f"{col}{start}:{col}{end}")

# 컬럼 폭 설정
widths = {"A": 22, "B": 30, "C": 30, "D": 10, "E": 14, "F": 50, "G": 50, "H": 10, "I": 14, "J": 50, "K": 50}
for col, w in widths.items(): ws.column_dimensions[col].width = w

wb.save(OUTFILE)
print(f"✨ 완료: {OUTFILE}")
import boto3
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ===== 설정 =====
PROFILE = "AIR-P"          # 없으면 None
REGION  = "ap-northeast-2"
OUTFILE = "security_groups_rules_side_by_side.xlsx"
# ===============

session = boto3.Session(profile_name=PROFILE) if PROFILE else boto3.Session()
ec2 = session.client("ec2", region_name=REGION)

# -------------------------
# VPC Name 맵 (Tag:Name)
# - Name 태그 없으면 VPC ID로 채워서 "-" 방지
# -------------------------
vpc_name_by_id = {}
for page in ec2.get_paginator("describe_vpcs").paginate():
    for v in page["Vpcs"]:
        vpc_id = v["VpcId"]
        name = None
        for t in (v.get("Tags") or []):
            if t.get("Key") == "Name" and t.get("Value"):
                name = t.get("Value")
                break
        vpc_name_by_id[vpc_id] = name or vpc_id

# -------------------------
# SG 전체 조회
# -------------------------
sgs = []
for page in ec2.get_paginator("describe_security_groups").paginate():
    sgs.extend(page["SecurityGroups"])

# SG 참조 표기용 (sg-xxxx(Name))
sg_name_by_id = {}
for sg in sgs:
    gid = sg.get("GroupId")
    gname = sg.get("GroupName")
    if gid:
        sg_name_by_id[gid] = gname or gid

# -------------------------
# 룰 확장
# - 소스가 여러 개면 "소스만" 행 분리
# - 비고에는 각 소스에 달린 Description 값을 기록
# 반환: [(Type, PortRange, Source, Remark), ...]
# -------------------------
def expand_rules(perms):
    out = []
    for p in (perms or []):
        proto = p.get("IpProtocol", "-1")
        proto = "all" if proto == "-1" else str(proto)

        fp, tp = p.get("FromPort"), p.get("ToPort")
        if fp is None or tp is None:
            pr = "-"
        else:
            pr = str(fp) if fp == tp else f"{fp}-{tp}"

        # (source, remark) 리스트
        src_items = []

        # IPv4 CIDR
        for r in (p.get("IpRanges") or []):
            cidr = r.get("CidrIp")
            if cidr:
                remark = r.get("Description") or "-"
                src_items.append((cidr, remark))

        # IPv6 CIDR
        for r in (p.get("Ipv6Ranges") or []):
            cidr6 = r.get("CidrIpv6")
            if cidr6:
                remark = r.get("Description") or "-"
                src_items.append((cidr6, remark))

        # SG reference
        for g in (p.get("UserIdGroupPairs") or []):
            gid = g.get("GroupId")
            if gid:
                ref_name = sg_name_by_id.get(gid, gid)
                src = f"{gid}({ref_name})"
                remark = g.get("Description") or "-"
                src_items.append((src, remark))

        # Prefix list
        for pl in (p.get("PrefixListIds") or []):
            plid = pl.get("PrefixListId")
            if plid:
                remark = pl.get("Description") or "-"
                src_items.append((plid, remark))

        # 소스가 없으면 1줄
        if not src_items:
            out.append((proto, pr, "-", "-"))
        else:
            for src, remark in src_items:
                out.append((proto, pr, src, remark))

    # 룰이 아예 없으면 1줄
    if not out:
        out.append(("-", "-", "-", "-"))

    return out

# -------------------------
# 엑셀 생성 (원래 포맷: Inbound/Outbound 옆으로 정렬)
# - A/B/C는 모든 행에 값 채움
# - SG 단위로 A/B/C 세로 병합
# -------------------------
wb = Workbook()
ws = wb.active
ws.title = "SecurityGroups"

# 스타일
header_fill = PatternFill("solid", fgColor="E6E6E6")
bold = Font(bold=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
top_wrap = Alignment(vertical="top", wrap_text=True)
thin = Side(style="thin", color="808080")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# 헤더 (2줄 + merge)
ws["A1"] = "VPC Name"
ws["B1"] = "Security Groups Name"
ws["C1"] = "Group ID"
ws["D1"] = "Inbound Rule"
ws["G1"] = "비고 (Inbound)"
ws["H1"] = "Outbound Rule"
ws["K1"] = "비고 (Outbound)"

ws["D2"] = "Type"
ws["E2"] = "Port Range"
ws["F2"] = "Source"
ws["H2"] = "Type"
ws["I2"] = "Port Range"
ws["J2"] = "Source"

ws.merge_cells("A1:A2")
ws.merge_cells("B1:B2")
ws.merge_cells("C1:C2")
ws.merge_cells("D1:F1")
ws.merge_cells("G1:G2")
ws.merge_cells("H1:J1")
ws.merge_cells("K1:K2")

ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 20
for r in (1, 2):
    for c in range(1, 12):
        cell = ws.cell(row=r, column=c)
        cell.fill = header_fill
        cell.font = bold
        cell.alignment = center
        cell.border = border

# 컬럼 폭
col_widths = {
    "A": 24,
    "B": 32,
    "C": 18,
    "D": 10,
    "E": 14,
    "F": 36,
    "G": 24,  # 비고 넓힘
    "H": 10,
    "I": 14,
    "J": 36,
    "K": 24,  # 비고 넓힘
}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

ws.freeze_panes = "A3"

# 데이터 작성 + SG 단위 병합
row_idx = 3
for sg in sorted(sgs, key=lambda x: (x.get("GroupName") or "").lower()):
    vpc_id = sg.get("VpcId")
    sg_id = sg.get("GroupId") or ""
    sg_name = (sg.get("GroupName") or sg_id or "")

    # VPC 없는 케이스도 "-" 대신 식별값
    vpc_name = vpc_name_by_id.get(vpc_id, vpc_id) if vpc_id else "NO_VPC"

    in_rules = expand_rules(sg.get("IpPermissions") or [])
    out_rules = expand_rules(sg.get("IpPermissionsEgress") or [])

    n = max(len(in_rules), len(out_rules))
    start_row = row_idx

    for i in range(n):
        it, ip, isrc, iremark = in_rules[i] if i < len(in_rules) else ("-", "-", "-", "-")
        ot, op, osrc, oremark = out_rules[i] if i < len(out_rules) else ("-", "-", "-", "-")

        values = [
            vpc_name, sg_name, sg_id,
            it, ip, isrc, iremark,
            ot, op, osrc, oremark
        ]

        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=c, value=v)
            cell.alignment = top_wrap
            cell.border = border

        row_idx += 1

    end_row = row_idx - 1

    # SG 단위로 A/B/C 세로 병합
    if end_row > start_row:
        ws.merge_cells(f"A{start_row}:A{end_row}")
        ws.merge_cells(f"B{start_row}:B{end_row}")
        ws.merge_cells(f"C{start_row}:C{end_row}")

        ws[f"A{start_row}"].alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        ws[f"B{start_row}"].alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        ws[f"C{start_row}"].alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

wb.save(OUTFILE)
print(f"Saved: {OUTFILE}")
